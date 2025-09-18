import requests
import threading
import pandas as pd
from openpyxl import load_workbook

class WeatherThread(threading.Thread):
    def __init__(self, district, result, row_number):
        super().__init__()
        self.district = district
        self.result = result
        self.row_number = row_number

    def run(self):
        weather_data = self.request_weather(self.district)
        self.result[self.district] = weather_data

    def request_weather(self, district):
        key = 'a9f3b3ad29e04f8eb8273440241903'
        api = 'http://api.weatherapi.com/v1/current.json?'
        payload = {
            "key": key,
            "q": district,
            "aqi": 'no'
        }
        try:
            response = requests.get(api, params=payload, timeout=1000)
            data = response.json()
            print("Data from API:", data)
            weather_info = {
                "Temperature (°C)": data['current']['temp_c'],
                "Humidity": data['current']['humidity'],
                "Condition": data['current']['condition']['text'],
                "Wind Speed (km/h)": data['current']['wind_kph'],
                "Wind Direction": data['current']['wind_dir'],
                "UV Index": data['current']['uv']
            }
            return weather_info
        except Exception as e:
            print(f"Failed to fetch weather data for {district}: {e}")
            return None

def main():
    cities = pd.read_excel("nusa_tenggara_barat.xlsx", index_col=0)
    results = {}
    threads = []

    for idx, district in enumerate(cities.index, start=2):
        thread = WeatherThread(district, results, idx)
        threads.append(thread)
        thread.start()

    for thread in threads:
        thread.join()

    wb = load_workbook("nusa_tenggara_barat.xlsx")
    ws = wb.active

    for district, weather_data in results.items():
        row_number = cities.index.get_loc(district) + 2
        if weather_data is not None: 
            for col, value in enumerate(weather_data.values(), start=2):
                ws.cell(row=row_number, column=col, value=value)

    wb.save("nusa_tenggara_barat.xlsx")

if __name__ == '__main__':
    main()
